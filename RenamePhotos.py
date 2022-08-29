import os
import argparse
import glob
import csv
from openpyxl import load_workbook


def get_args():
    '''Get arguements from user'''
    parser = argparse.ArgumentParser(description='STAIR Generic File Renamer')
    parser.add_argument(
        '-d', '--directory', metavar='directory', default=os.path.dirname(__file__), help='The directory that has the items to be renamed.'
    )
    parser.add_argument(
        '-r', '--reverse', action='store_true', help='Reverse the order of the renaming.'
    )
    return parser.parse_args()


def countWork(workingDir):
    '''Enumerate the possible filetypes in our work folder.'''
    osSep = os.sep

    jpgList = glob.glob(workingDir + osSep + '*.jpg')
    nefList = glob.glob(workingDir + osSep + '.nef')
    csvList = glob.glob(workingDir + osSep + '*.csv')
    xlsxList = glob.glob(workingDir + osSep + '*.xlsx')

    return jpgList, nefList, csvList, xlsxList


def readCsv(csvFile):
    '''Read in a csv that has filename pairs and return them as a dictionary filename pairs'''
    filenameDict = {}

    for file in csvFile:
        dataFileNames = csv.reader(open(file, newline=''))
        for sourceFilename, destFilename in dataFileNames:
            filenameDict[sourceFilename] = destFilename

    return filenameDict


def readXlsx(xlsxList):
    '''Read in a list of formatted XLSX docs and read the contents and return it as a dict of filename pairs'''
    filenameDict = {}

    for file in xlsxList:

        wb = load_workbook(filename=file)
        # Raw Entry is the sheet that we are targetting to read all of our data
        ws = wb['Raw Entry']
        # This cell is where we find the correct prefix for the files coming out of the cameras 
        prefixVal = ws['E2'].value[-4:-1]

        # 
        positionNum = 1
        storedVal = 0

        for destVal, srcVal, formatSrc, formatDest, photog, sale in ws.values:
            try:
                if int(storedVal) != int(destVal):
                    positionNum = 1
                else:
                    positionNum = positionNum + 1
            except ValueError:
                pass
            except TypeError:
                pass

            if destVal is None:
                break
            else:
                srcVal = str(srcVal)
                destVal = str(destVal) + '_' + str(positionNum) + '.jpg'
                srcVal = prefixVal + '_' + srcVal.zfill(4) + '.jpg'
                filenameDict[srcVal] = destVal
                storedVal = destVal[:7]

    return filenameDict


def renameFiles(workingDir, csvFilenames, xlsxFilenames, reverse):

    fileNames = csvFilenames | xlsxFilenames

    osSep = os.sep

    for filename in fileNames:
        if (reverse):
            try:
                os.rename(workingDir + osSep + fileNames[filename],
                          workingDir + osSep + filename)
                print(f'''Filename: {fileNames[filename]} ==> {filename}''')
            except FileNotFoundError:
                print(f'''File Not Found: {filename}''')
        else:
            try:
                os.rename(workingDir + osSep + filename,
                          workingDir + osSep + fileNames[filename])
                print(f'''Filename: {filename} ==> {fileNames[filename]}''')
            except FileNotFoundError:
                print(f'''File Not Found: {fileNames[filename]}''')
            except TypeError:
                continue
            except FileExistsError:
                print(f'{fileNames[filename]} already exists.')


def main():
    args = get_args()

    jpgList, nefList, csvList, xlsxList = countWork(args.directory)

    renameFiles(args.directory, readCsv(csvList),
                readXlsx(xlsxList), args.reverse)

    print(f'''
    Current working directory {args.directory}
    Discovered {len(jpgList)} jpgs.
    Discovered {len(nefList)} Nikon RAW files.
    Discovered {len(csvList)} CSV files.
    Discovered {len(xlsxList)} Excel spreadsheets.
    ''')


if __name__ == '__main__':
    main()
